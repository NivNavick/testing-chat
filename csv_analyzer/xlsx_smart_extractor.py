"""
Smart XLSX Extractor - Language-agnostic automatic table detection.

Uses STRUCTURAL patterns (not keywords) to detect:
- Header rows: first non-empty row near the top
- Data rows: similar structure to header
- Table boundaries: structural changes

Works with ANY language (Hebrew, English, Arabic, Chinese, etc.)
"""

import logging
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class DetectedTable:
    """A detected table with its boundaries."""
    start_row: int  # Header row (1-indexed)
    end_row: int    # Last data row (1-indexed)
    header_values: List[str]
    data_row_count: int
    confidence: float
    detection_method: str


def get_row_values(ws: Worksheet, row_idx: int, max_col: int) -> List[Any]:
    """Get values for a row."""
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def count_filled(row_values: List[Any]) -> int:
    """Count non-empty cells."""
    return sum(1 for v in row_values if v is not None and str(v).strip())


def get_filled_positions(row_values: List[Any]) -> set:
    """Get positions of filled cells."""
    return {i for i, v in enumerate(row_values) if v is not None and str(v).strip()}


def row_structure_similarity(row1_positions: set, row2_positions: set) -> float:
    """Calculate structural similarity between two rows (0 to 1)."""
    if not row1_positions or not row2_positions:
        return 0.0
    intersection = len(row1_positions & row2_positions)
    union = len(row1_positions | row2_positions)
    return intersection / union if union > 0 else 0.0


def has_formula(row_values: List[Any]) -> bool:
    """Check if any cell contains a formula."""
    for v in row_values:
        if v is not None and str(v).strip().startswith('='):
            return True
    return False


def find_first_data_row(ws: Worksheet, max_col: int, max_scan: int = 20) -> Optional[int]:
    """
    Find the first row that looks like it has data.
    
    Strategy: Find the first row with at least 3 filled cells.
    """
    for row_idx in range(1, min(max_scan, ws.max_row) + 1):
        row_values = get_row_values(ws, row_idx, max_col)
        filled = count_filled(row_values)
        if filled >= 3:
            return row_idx
    return None


def find_table_end(
    ws: Worksheet, 
    header_row: int, 
    header_positions: set,
    max_col: int,
    max_row: int,
    min_similarity: float = 0.25
) -> int:
    """
    Find where the table ends.
    
    Strategy: 
    - Count consecutive rows that match header structure
    - Stop at blank rows or structural changes
    """
    last_data_row = header_row
    blank_streak = 0
    
    for row_idx in range(header_row + 1, max_row + 1):
        row_values = get_row_values(ws, row_idx, max_col)
        filled = count_filled(row_values)
        
        # Empty row
        if filled == 0:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        
        # Formula row (summary)
        if has_formula(row_values):
            break
        
        blank_streak = 0
        
        # Check structural similarity to header
        row_positions = get_filled_positions(row_values)
        similarity = row_structure_similarity(header_positions, row_positions)
        
        # First 2 columns must have values (key identifiers)
        first_two = [row_values[i] for i in range(min(2, len(row_values)))]
        has_keys = all(v is not None and str(v).strip() for v in first_two)
        
        if similarity >= min_similarity and has_keys:
            last_data_row = row_idx
        else:
            # Structure changed significantly - likely end of table
            break
    
    return last_data_row


def detect_main_table(file_path: str, sheet_name: str = None) -> Optional[DetectedTable]:
    """
    Automatically detect the main data table in an Excel sheet.
    
    Simple strategy:
    1. Find first non-empty row (header)
    2. Find where data stops (structure change or blank rows)
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    
    max_row = min(ws.max_row, 500)
    max_col = min(ws.max_column, 50)
    
    logger.info(f"Analyzing sheet: {ws.title} ({max_row} rows, {max_col} cols)")
    
    # Step 1: Find the header (first row with data)
    header_row = find_first_data_row(ws, max_col)
    
    if header_row is None:
        logger.warning("No header row found")
        wb.close()
        return None
    
    header_values = get_row_values(ws, header_row, max_col)
    header_positions = get_filled_positions(header_values)
    
    logger.info(f"Header found at row {header_row}: {[v for v in header_values[:5] if v]}")
    
    # Step 2: Find table end
    end_row = find_table_end(
        ws, header_row, header_positions, 
        max_col, max_row
    )
    
    data_row_count = end_row - header_row
    
    if data_row_count < 1:
        logger.warning("No data rows found")
        wb.close()
        return None
    
    wb.close()
    
    table = DetectedTable(
        start_row=header_row,
        end_row=end_row,
        header_values=[str(v) if v else '' for v in header_values],
        data_row_count=data_row_count,
        confidence=min(0.95, 0.6 + min(data_row_count / 100, 0.35)),
        detection_method="first_row_structural"
    )
    
    logger.info(
        f"Detected table: rows {table.start_row}-{table.end_row} "
        f"({table.data_row_count} data rows, {table.confidence:.0%} confidence)"
    )
    
    return table


def extract_table(file_path: str, sheet_name: str = None) -> Optional[pd.DataFrame]:
    """
    Extract the main table from an Excel file.
    
    Automatically detects table boundaries using structural analysis.
    """
    table = detect_main_table(file_path, sheet_name)
    
    if table is None:
        logger.warning("Could not detect table, reading entire sheet")
        return pd.read_excel(file_path, sheet_name=sheet_name or 0)
    
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name or 0,
        skiprows=table.start_row - 1,
        nrows=table.data_row_count,
        header=0
    )
    
    logger.info(f"Extracted {len(df)} rows from detected table")
    return df


def extract_all_tables(file_path: str, sheet_name: str = None) -> List[Tuple[DetectedTable, pd.DataFrame]]:
    """
    Extract all tables from an Excel sheet.
    
    Finds multiple tables by scanning for header-like rows after gaps.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    
    max_row = min(ws.max_row, 500)
    max_col = min(ws.max_column, 50)
    
    tables = []
    current_row = 1
    
    while current_row < max_row:
        # Find next header
        header_row = None
        for row_idx in range(current_row, max_row + 1):
            row_values = get_row_values(ws, row_idx, max_col)
            if count_filled(row_values) >= 3:
                header_row = row_idx
                break
        
        if header_row is None:
            break
        
        header_values = get_row_values(ws, header_row, max_col)
        header_positions = get_filled_positions(header_values)
        
        # Find table end
        end_row = find_table_end(
            ws, header_row, header_positions, 
            max_col, max_row
        )
        
        data_row_count = end_row - header_row
        
        if data_row_count >= 2:  # At least 2 data rows
            table = DetectedTable(
                start_row=header_row,
                end_row=end_row,
                header_values=[str(v) if v else '' for v in header_values[:10]],
                data_row_count=data_row_count,
                confidence=0.8,
                detection_method="multi_table_scan"
            )
            
            try:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name or 0,
                    skiprows=header_row - 1,
                    nrows=data_row_count,
                    header=0
                )
                tables.append((table, df))
                logger.info(f"Found table at rows {header_row}-{end_row} ({data_row_count} rows)")
            except Exception as e:
                logger.warning(f"Could not read table at row {header_row}: {e}")
        
        # Move to after this table (skip blank rows)
        current_row = end_row + 2
    
    wb.close()
    return tables


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python xlsx_smart_extractor.py <file.xlsx>")
        sys.exit(1)
    
    logging.basicConfig(level=logging.INFO)
    
    file_path = sys.argv[1]
    
    print("\n=== Detecting Main Table ===\n")
    table = detect_main_table(file_path)
    
    if table:
        print(f"Main table: rows {table.start_row}-{table.end_row}")
        print(f"Data rows: {table.data_row_count}")
        print(f"Confidence: {table.confidence:.0%}")
        print(f"Method: {table.detection_method}")
        print(f"Headers: {[h for h in table.header_values[:5] if h]}...")
        
        # Extract and show preview
        df = extract_table(file_path)
        print(f"\nShape: {df.shape}")
        print(f"\nFirst 3 rows:")
        print(df.head(3).to_string())
        print(f"\nLast 3 rows:")
        print(df.tail(3).to_string())
