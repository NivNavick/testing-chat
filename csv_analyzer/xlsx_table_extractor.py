#!/usr/bin/env python3
"""
XLSX Table Extractor - Extract multiple tables from a single Excel sheet.

This script detects and extracts multiple logical tables that exist within
a single Excel sheet. It handles various patterns:
- Tables separated by blank rows
- Tables with different column structures
- Section headers and merged cells
- Mixed Hebrew/English content

Usage:
    # Extract all tables and preview them
    python xlsx_table_extractor.py --file data.xlsx --preview
    
    # Export each table to separate CSV files
    python xlsx_table_extractor.py --file data.xlsx --output-dir output/
    
    # Get detailed analysis with cell formatting info
    python xlsx_table_extractor.py --file data.xlsx --analyze
    
    # Specify a particular sheet
    python xlsx_table_extractor.py --file data.xlsx --sheet "Sheet2"
"""

import argparse
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


@dataclass
class TableRegion:
    """Represents a detected table region in the sheet."""
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    header_row: int
    has_header: bool
    cell_count: int
    confidence: float  # How confident we are this is a real table
    
    def __str__(self):
        return (
            f"Table[rows {self.start_row}-{self.end_row}, "
            f"cols {self.start_col}-{self.end_col}, "
            f"confidence {self.confidence:.0%}]"
        )


@dataclass
class ExtractedTable:
    """Represents an extracted table with its data and metadata."""
    region: TableRegion
    dataframe: pd.DataFrame
    title: Optional[str] = None  # If there's a title row above the table
    sheet_name: str = ""
    
    def to_csv(self, output_path: Path):
        """Export table to CSV."""
        self.dataframe.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"Exported table to {output_path}")


class XLSXTableExtractor:
    """Extract multiple tables from an XLSX sheet."""
    
    def __init__(
        self,
        min_rows: int = 2,
        min_cols: int = 2,
        blank_row_threshold: int = 2,
        min_confidence: float = 0.3,
        mode: str = "auto",
        max_gap: int = 3
    ):
        """
        Initialize extractor.
        
        Args:
            min_rows: Minimum rows for a valid table (including header)
            min_cols: Minimum columns for a valid table
            blank_row_threshold: Number of consecutive blank rows to split tables
            min_confidence: Minimum confidence score to consider a region as a table
            mode: Detection mode - "auto" (split on blanks), "single" (one table), "smart" (detect structure)
            max_gap: Maximum blank rows allowed within a table in smart mode (default: 3)
        """
        self.min_rows = min_rows
        self.min_cols = min_cols
        self.blank_row_threshold = blank_row_threshold
        self.min_confidence = min_confidence
        self.mode = mode
        self.max_gap = max_gap
    
    def extract_tables(
        self,
        file_path: Path,
        sheet_name: Optional[str] = None
    ) -> List[ExtractedTable]:
        """
        Extract all tables from the specified sheet.
        
        Args:
            file_path: Path to XLSX file
            sheet_name: Sheet name or index (default: first sheet)
        
        Returns:
            List of ExtractedTable objects
        """
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        
        sheet_name_str = ws.title
        logger.info(f"Processing sheet: {sheet_name_str}")
        
        # Detect table regions
        regions = self._detect_table_regions(ws)
        logger.info(f"Detected {len(regions)} potential table(s)")
        
        # Extract data from each region
        tables = []
        for i, region in enumerate(regions, 1):
            logger.info(f"Extracting table {i}: {region}")
            
            # Look for title row above the table
            title = self._extract_title(ws, region)
            
            # Extract the table data
            df = self._extract_dataframe(ws, region)
            
            if df is not None and not df.empty:
                table = ExtractedTable(
                    region=region,
                    dataframe=df,
                    title=title,
                    sheet_name=sheet_name_str
                )
                tables.append(table)
                logger.info(f"  ✓ Extracted table with shape {df.shape}")
            else:
                logger.warning(f"  ✗ Failed to extract data from region")
        
        wb.close()
        return tables
    
    def _detect_table_regions(self, ws: Worksheet) -> List[TableRegion]:
        """Detect all table regions in the worksheet."""
        if self.mode == "single":
            return self._detect_single_table(ws)
        elif self.mode == "smart":
            return self._detect_smart_tables(ws)
        else:  # auto mode
            return self._detect_auto_tables(ws)
    
    def _detect_single_table(self, ws: Worksheet) -> List[TableRegion]:
        """Treat entire non-empty region as one table."""
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Find bounding box of all data
        min_row = None
        max_data_row = None
        min_col = None
        max_data_col = None
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if self._is_cell_filled(cell):
                    if min_row is None:
                        min_row = row
                    max_data_row = row
                    if min_col is None or col < min_col:
                        min_col = col
                    if max_data_col is None or col > max_data_col:
                        max_data_col = col
        
        if min_row is None or (max_data_row - min_row + 1) < self.min_rows:
            return []
        
        region = TableRegion(
            start_row=min_row,
            end_row=max_data_row,
            start_col=min_col or 1,
            end_col=max_data_col or max_col,
            header_row=min_row,
            has_header=True,
            cell_count=(max_data_row - min_row + 1) * (max_data_col - min_col + 1) if min_col and max_data_col else 0,
            confidence=1.0
        )
        
        return [region]
    
    def _detect_smart_tables(self, ws: Worksheet) -> List[TableRegion]:
        """Smart detection: find continuous data regions, ignore sparse blank rows."""
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Build row density map (how many cells are filled per row)
        row_density = {}
        for row in range(1, max_row + 1):
            filled_count = 0
            for col in range(1, max_col + 1):
                if self._is_cell_filled(ws.cell(row, col)):
                    filled_count += 1
            if filled_count > 0:
                row_density[row] = filled_count
        
        if not row_density:
            return []
        
        # Find the main data block (continuous rows with similar density)
        rows = sorted(row_density.keys())
        
        # Find largest continuous block
        regions = []
        current_block_start = None
        prev_row = None
        max_gap = self.max_gap  # Use configurable gap parameter
        
        for row in rows:
            if current_block_start is None:
                current_block_start = row
                prev_row = row
            elif row - prev_row <= max_gap:
                # Continue current block
                prev_row = row
            else:
                # Gap too large, end current block
                if prev_row - current_block_start + 1 >= self.min_rows:
                    regions.append((current_block_start, prev_row))
                current_block_start = row
                prev_row = row
        
        # Don't forget last block
        if current_block_start and prev_row:
            if prev_row - current_block_start + 1 >= self.min_rows:
                regions.append((current_block_start, prev_row))
        
        # Convert to TableRegion objects
        table_regions = []
        for start_row, end_row in regions:
            # Find column bounds for this region
            min_col = max_col
            max_data_col = 1
            for row in range(start_row, end_row + 1):
                for col in range(1, max_col + 1):
                    if self._is_cell_filled(ws.cell(row, col)):
                        min_col = min(min_col, col)
                        max_data_col = max(max_data_col, col)
            
            region = TableRegion(
                start_row=start_row,
                end_row=end_row,
                start_col=min_col,
                end_col=max_data_col,
                header_row=start_row,
                has_header=True,
                cell_count=(end_row - start_row + 1) * (max_data_col - min_col + 1),
                confidence=0.9
            )
            table_regions.append(region)
        
        return table_regions
    
    def _detect_auto_tables(self, ws: Worksheet) -> List[TableRegion]:
        """Original auto detection method."""
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Build a map of non-empty cells
        cell_map = {}
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if self._is_cell_filled(cell):
                    cell_map[(row, col)] = cell
        
        if not cell_map:
            return []
        
        # Find contiguous regions of data
        regions = []
        visited = set()
        
        # Scan row by row
        current_region = None
        blank_row_count = 0
        
        for row in range(1, max_row + 1):
            row_cells = [(r, c) for (r, c) in cell_map.keys() if r == row]
            
            if not row_cells:
                # Blank row
                blank_row_count += 1
                
                # If we hit enough blank rows, end the current region
                if current_region and blank_row_count >= self.blank_row_threshold:
                    region = self._finalize_region(current_region, ws)
                    if region and region.confidence >= self.min_confidence:
                        regions.append(region)
                    current_region = None
                    blank_row_count = 0
            else:
                # Non-blank row
                blank_row_count = 0
                
                cols = [c for (r, c) in row_cells]
                min_col, max_col_in_row = min(cols), max(cols)
                
                if current_region is None:
                    # Start new region
                    current_region = {
                        'start_row': row,
                        'end_row': row,
                        'start_col': min_col,
                        'end_col': max_col_in_row,
                        'cells': row_cells
                    }
                else:
                    # Check if this row belongs to current region
                    # (similar column structure)
                    col_overlap = (
                        min_col <= current_region['end_col'] and
                        max_col_in_row >= current_region['start_col']
                    )
                    
                    col_diff = abs(len(cols) - (current_region['end_col'] - current_region['start_col'] + 1))
                    
                    # If columns are very different, might be a new table
                    if col_overlap and col_diff < 5:
                        # Extend current region
                        current_region['end_row'] = row
                        current_region['start_col'] = min(current_region['start_col'], min_col)
                        current_region['end_col'] = max(current_region['end_col'], max_col_in_row)
                        current_region['cells'].extend(row_cells)
                    else:
                        # Column structure changed significantly - new table
                        region = self._finalize_region(current_region, ws)
                        if region and region.confidence >= self.min_confidence:
                            regions.append(region)
                        
                        # Start new region
                        current_region = {
                            'start_row': row,
                            'end_row': row,
                            'start_col': min_col,
                            'end_col': max_col_in_row,
                            'cells': row_cells
                        }
        
        # Don't forget the last region
        if current_region:
            region = self._finalize_region(current_region, ws)
            if region and region.confidence >= self.min_confidence:
                regions.append(region)
        
        return regions
    
    def _finalize_region(self, region_dict: Dict, ws: Worksheet) -> Optional[TableRegion]:
        """Convert region dict to TableRegion with validation."""
        row_count = region_dict['end_row'] - region_dict['start_row'] + 1
        col_count = region_dict['end_col'] - region_dict['start_col'] + 1
        
        # Check minimum size
        if row_count < self.min_rows or col_count < self.min_cols:
            return None
        
        # Detect header row (usually first row, or one with different formatting)
        header_row = region_dict['start_row']
        has_header = True
        
        first_row_cells = [
            ws.cell(region_dict['start_row'], c)
            for c in range(region_dict['start_col'], region_dict['end_col'] + 1)
        ]
        
        # Check if first row looks like a header (bold, different style, text content)
        bold_count = sum(1 for cell in first_row_cells if cell.font and cell.font.bold)
        if bold_count > len(first_row_cells) * 0.5:
            has_header = True
        
        # Calculate confidence score
        confidence = self._calculate_confidence(region_dict, ws)
        
        return TableRegion(
            start_row=region_dict['start_row'],
            end_row=region_dict['end_row'],
            start_col=region_dict['start_col'],
            end_col=region_dict['end_col'],
            header_row=header_row,
            has_header=has_header,
            cell_count=len(region_dict['cells']),
            confidence=confidence
        )
    
    def _calculate_confidence(self, region_dict: Dict, ws: Worksheet) -> float:
        """Calculate confidence that this region is a real table."""
        row_count = region_dict['end_row'] - region_dict['start_row'] + 1
        col_count = region_dict['end_col'] - region_dict['start_col'] + 1
        expected_cells = row_count * col_count
        actual_cells = len(region_dict['cells'])
        
        # Density: how full is the region?
        density = actual_cells / expected_cells if expected_cells > 0 else 0
        
        # Size bonus: bigger tables are more likely real
        size_score = min(1.0, (row_count * col_count) / 20)
        
        # Header check: does first row look different?
        first_row = region_dict['start_row']
        first_row_cells = [
            ws.cell(first_row, c)
            for c in range(region_dict['start_col'], region_dict['end_col'] + 1)
        ]
        bold_in_header = sum(1 for cell in first_row_cells if cell.font and cell.font.bold)
        header_score = min(1.0, bold_in_header / len(first_row_cells)) if first_row_cells else 0
        
        # Combined confidence
        confidence = (density * 0.5 + size_score * 0.3 + header_score * 0.2)
        
        return confidence
    
    def _extract_title(self, ws: Worksheet, region: TableRegion) -> Optional[str]:
        """Try to find a title row above the table."""
        if region.start_row <= 1:
            return None
        
        # Check 1-3 rows above
        for offset in range(1, min(4, region.start_row)):
            row = region.start_row - offset
            
            # Check if this row has a single merged cell or centered text
            cell = ws.cell(row, region.start_col)
            if cell.value and isinstance(cell.value, str):
                # Check if this looks like a title (short text, possibly bold)
                if len(str(cell.value).strip()) > 0:
                    return str(cell.value).strip()
        
        return None
    
    def _extract_dataframe(self, ws: Worksheet, region: TableRegion) -> Optional[pd.DataFrame]:
        """Extract data from region as a pandas DataFrame."""
        try:
            # Extract all cell values
            data = []
            for row in range(region.start_row, region.end_row + 1):
                row_data = []
                for col in range(region.start_col, region.end_col + 1):
                    cell = ws.cell(row, col)
                    row_data.append(cell.value)
                data.append(row_data)
            
            if not data:
                return None
            
            # Convert to DataFrame
            if region.has_header:
                # First row is header
                df = pd.DataFrame(data[1:], columns=data[0])
            else:
                # No header, use column numbers
                df = pd.DataFrame(data)
            
            # Clean up
            df = df.dropna(how='all', axis=0)  # Remove completely empty rows
            df = df.dropna(how='all', axis=1)  # Remove completely empty columns
            
            # Clean column names
            df.columns = [self._clean_column_name(col) for col in df.columns]
            
            return df
        
        except Exception as e:
            logger.error(f"Error extracting dataframe: {e}")
            return None
    
    def _clean_column_name(self, name) -> str:
        """Clean column name."""
        if name is None:
            return "Unnamed"
        
        name = str(name).strip()
        if not name:
            return "Unnamed"
        
        return name
    
    def _is_cell_filled(self, cell: Cell) -> bool:
        """Check if a cell has content."""
        if cell.value is None:
            return False
        
        # Check if it's whitespace
        if isinstance(cell.value, str) and not cell.value.strip():
            return False
        
        return True


def preview_tables(tables: List[ExtractedTable]):
    """Print preview of extracted tables."""
    print("\n" + "="*80)
    print(f"EXTRACTED {len(tables)} TABLE(S)")
    print("="*80)
    
    for i, table in enumerate(tables, 1):
        print(f"\n{'─'*80}")
        print(f"TABLE {i}: {table.region}")
        print(f"Sheet: {table.sheet_name}")
        if table.title:
            print(f"Title: {table.title}")
        print(f"Shape: {table.dataframe.shape[0]} rows × {table.dataframe.shape[1]} columns")
        print(f"Columns: {', '.join(table.dataframe.columns.tolist())}")
        print(f"\nFirst 5 rows:")
        print(table.dataframe.head(5).to_string(index=False))
    
    print("\n" + "="*80)


def analyze_tables(tables: List[ExtractedTable]):
    """Provide detailed analysis of extracted tables."""
    print("\n" + "="*80)
    print(f"DETAILED ANALYSIS OF {len(tables)} TABLE(S)")
    print("="*80)
    
    for i, table in enumerate(tables, 1):
        df = table.dataframe
        region = table.region
        
        print(f"\n{'─'*80}")
        print(f"TABLE {i}")
        print(f"{'─'*80}")
        print(f"Location: Rows {region.start_row}-{region.end_row}, "
              f"Cols {region.start_col}-{region.end_col}")
        print(f"Confidence: {region.confidence:.1%}")
        print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
        
        if table.title:
            print(f"Title: {table.title}")
        
        print(f"\nColumns:")
        for idx, col in enumerate(df.columns):
            try:
                # Handle potential duplicate column names
                col_data = df.iloc[:, idx]
                dtype = col_data.dtype
                non_null = col_data.notna().sum()
                null_pct = (col_data.isna().sum() / len(df)) * 100 if len(df) > 0 else 0
                
                # Sample values
                samples = col_data.dropna().head(3).tolist()
                samples_str = ", ".join(str(s) for s in samples)
                if len(samples_str) > 50:
                    samples_str = samples_str[:47] + "..."
                
                print(f"  • {col}")
                print(f"    Type: {dtype}, Non-null: {non_null}/{len(df)} ({100-null_pct:.0f}%)")
                print(f"    Samples: {samples_str}")
            except Exception as e:
                print(f"  • {col}")
                print(f"    Error analyzing column: {e}")
        
        print(f"\nData Preview:")
        print(df.head(3).to_string(index=False))
    
    print("\n" + "="*80)


def export_tables(tables: List[ExtractedTable], output_dir: Path, base_name: str):
    """Export tables to separate CSV files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Exporting {len(tables)} table(s) to {output_dir}")
    
    for i, table in enumerate(tables, 1):
        # Create filename
        if table.title:
            # Use title in filename
            safe_title = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' 
                                for c in table.title)
            safe_title = safe_title.strip().replace(' ', '_')
            filename = f"{base_name}_table{i}_{safe_title}.csv"
        else:
            filename = f"{base_name}_table{i}.csv"
        
        output_path = output_dir / filename
        table.to_csv(output_path)
        print(f"  ✓ Table {i}: {output_path.name}")
    
    print(f"\n✅ Export complete!")


def main():
    parser = argparse.ArgumentParser(
        description="Extract multiple tables from an XLSX file"
    )
    
    parser.add_argument(
        "--file",
        type=str,
        required=True,
        help="Path to XLSX file"
    )
    parser.add_argument(
        "--sheet",
        type=str,
        help="Sheet name (default: first sheet)"
    )
    parser.add_argument(
        "--preview",
        action="store_true",
        help="Preview extracted tables"
    )
    parser.add_argument(
        "--analyze",
        action="store_true",
        help="Show detailed analysis of extracted tables"
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        help="Export each table to separate CSV files in this directory"
    )
    parser.add_argument(
        "--min-rows",
        type=int,
        default=2,
        help="Minimum rows for a valid table (default: 2)"
    )
    parser.add_argument(
        "--min-cols",
        type=int,
        default=2,
        help="Minimum columns for a valid table (default: 2)"
    )
    parser.add_argument(
        "--blank-threshold",
        type=int,
        default=2,
        help="Number of blank rows to split tables (default: 2)"
    )
    parser.add_argument(
        "--min-confidence",
        type=float,
        default=0.3,
        help="Minimum confidence score (0-1) to consider as table (default: 0.3)"
    )
    parser.add_argument(
        "--mode",
        type=str,
        choices=["auto", "single", "smart"],
        default="smart",
        help="Detection mode: 'auto' (split on blank rows), 'single' (treat as one table), 'smart' (find continuous blocks, default)"
    )
    parser.add_argument(
        "--max-gap",
        type=int,
        default=3,
        help="Maximum blank rows allowed within a table in smart mode (default: 3, use 1 for stricter splitting)"
    )
    
    args = parser.parse_args()
    
    # Validate file
    file_path = Path(args.file)
    if not file_path.exists():
        logger.error(f"File not found: {args.file}")
        return 1
    
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        logger.error(f"File must be .xlsx or .xlsm format")
        return 1
    
    # Create extractor
    extractor = XLSXTableExtractor(
        min_rows=args.min_rows,
        min_cols=args.min_cols,
        blank_row_threshold=args.blank_threshold,
        min_confidence=args.min_confidence,
        mode=args.mode,
        max_gap=args.max_gap
    )
    
    # Extract tables
    logger.info(f"Extracting tables from: {file_path}")
    tables = extractor.extract_tables(file_path, sheet_name=args.sheet)
    
    if not tables:
        logger.warning("No tables found in the sheet")
        return 1
    
    logger.info(f"Successfully extracted {len(tables)} table(s)")
    
    # Show results based on flags
    if args.analyze:
        analyze_tables(tables)
    elif args.preview:
        preview_tables(tables)
    
    # Export if requested
    if args.output_dir:
        base_name = file_path.stem
        output_dir = Path(args.output_dir)
        export_tables(tables, output_dir, base_name)
    
    # If no output flags specified, do a simple preview
    if not (args.preview or args.analyze or args.output_dir):
        preview_tables(tables)
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())

